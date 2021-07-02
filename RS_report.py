import openpyxl
from openpyxl import Workbook
def append_to_cell(location,text,sheet):
    sheet = wb.active
    x = sheet[location]
    #if not none
    if x.value:
         
         
         x.value=new
         return 
    else:
         x.value = text
         return     
wb = openpyxl.load_workbook('F:\\automation\Daily_report.xlsx')
sheet = wb.active
for x in range(12, 61):
     cell = "E%d" % x
     x1 = sheet[cell].value
     wb1 = openpyxl.load_workbook(r'F:\\automation\\kk.xlsx')
     sheet = wb1.active
     for z in range(4, 52):
         cell2 = "G%d" % z
         x2 = sheet[cell2].value
         if x1 == x2:
             cell3 = "P%d" % z
             x3 = sheet[cell3].value
             #print(x3)  
             cell4 = "S%d" % z
             x4 = sheet[cell4].value             
             sheet = wb.active
             cell5 = "F%d" % x 
             y = sheet[cell5]
             y.value = x3
             cell6 = "G%d" % x 
             y2 = sheet[cell6]
             y2.value = x4                 
             wb.save('F:\\automation\Daily_report.xlsx')                     
     wb3 = openpyxl.load_workbook('F:\\automation\ll.xlsx')
     sheet = wb3.active
     a = 4
     while True:
        cell7 = "L%d" % a
        a1 = sheet[cell7].value
        cell8 = "H%d" % a
        a2 = sheet[cell8].value    
        cell9 = "F%d" % a
        a3 = sheet[cell9].value 
        print(a1)
        print(a2)
        if a1 >= 80:
            if x1 == a3:
                 text = "Disk Usage more than 80%%:%s:%s%%" %(a2,a1)
                 cell10 = "H%d" % x
                 append_to_cell(cell10,text,sheet)
                 wb.save('F:\\automation\Daily_report.xlsx') 
        else:
             print("not ok") 
             break
        a = a + 1 
     sheet = wb.active
    #print(x1)    


wb.save('F:\\automation\Daily_report.xlsx')    
